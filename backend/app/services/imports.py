"""Import pipeline (T13): parse a standardized .xlsx or .csv against the fixed
template (one flat sheet; amounts always positive; columns date, type, amount,
wallet, source wallet, destination wallet, category, description, location),
resolve each row against the Account, flag duplicates against the database, and
insert the confirmed rows in one transaction.

Parsing is pure (no database access). Resolution maps the file's names to the
Account's ids — unknown or frozen names are import errors here — and the
CONTEXT.md rules themselves are enforced by `services.transactions`
(validate_create in the preview, create_transaction at confirm), so a row an
import writes behaves exactly like one typed into the form and the two paths
cannot drift. Only the template-shape guards (which columns a row of a given
Type may fill) stay here, because the service reasons in ids, not template
columns.
"""

from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any
import csv
import io
import re

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.dates import from_rome_day
from app.models import Category, Transaction, TransactionType, Wallet
from app.schemas import ImportRow, ImportRowInput, fmt_coord
from app.services import scoping, transactions as transaction_service

# Matches the Numeric(12, 2) column and TransactionCreate._MAX_AMOUNT.
_MAX_AMOUNT = Decimal("9999999999.99")

# The fixed template's columns (US 7.1): normalized header -> canonical field.
_COLUMNS = {
    "date": "date",
    "type": "type",
    "amount": "amount",
    "wallet": "wallet",
    "sourcewallet": "source_wallet",
    "destinationwallet": "destination_wallet",
    "category": "category",
    "description": "description",
    "location": "location",
}


class UnsupportedFile(Exception):
    """The upload is not a .csv or .xlsx file."""


class BadTemplate(Exception):
    """The file is not the fixed template: a required column is missing."""


class ImportValidationError(Exception):
    """A row violates a rule from CONTEXT.md; nothing is written."""


@dataclass
class RawRow:
    """One data row as parsed from the file: the extracted fields plus a parse
    error when a cell could not be read. Names are kept as written (resolution
    happens later, against the Account)."""

    number: int
    error: str | None = None
    type: str | None = None
    date: str | None = None
    amount: Decimal | None = None
    wallet: str | None = None
    source_wallet: str | None = None
    destination_wallet: str | None = None
    category: str | None = None
    description: str | None = None
    latitude: str | None = None
    longitude: str | None = None


# --- Parsing: pure, no database access ---------------------------------------


def parse_file(filename: str, content: bytes) -> list[RawRow]:
    """Parse an uploaded file against the template into raw rows. Raises
    UnsupportedFile for a wrong extension and BadTemplate when a required
    column is missing."""
    name = filename.lower()
    if name.endswith(".csv"):
        return _rows_to_raw(_csv_cells(content))
    if name.endswith(".xlsx"):
        return _rows_to_raw(_xlsx_cells(content))
    raise UnsupportedFile("Only .csv and .xlsx files are supported")


def _csv_cells(content: bytes) -> list[list[str]]:
    text = _decode_csv(content)
    delimiter = _sniff_delimiter(text)
    return [
        row
        for row in csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
        if any(cell.strip() for cell in row)
    ]


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _sniff_delimiter(text: str) -> str:
    """The template's cells carry no commas; a European Excel export separates
    columns with ';' instead. The header row tells them apart."""
    first_line = text.splitlines()[0] if text.splitlines() else ""
    if ";" in first_line and "," not in first_line:
        return ";"
    return ","


def _xlsx_cells(content: bytes) -> list[list[str]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        assert sheet is not None  # an .xlsx always has at least one sheet
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [_cell_text(cell) for cell in row]
            if any(cell.strip() for cell in cells):
                rows.append(cells)
        return rows
    finally:
        workbook.close()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _normalize_header(value: str) -> str:
    """"source wallet" and "source_wallet" are the same column."""
    return re.sub(r"[\s_]+", "", value.strip().lower())


def _rows_to_raw(cells: list[list[str]]) -> list[RawRow]:
    """Map the sheet's rows onto the template and parse each data row."""
    if not cells:
        return []
    field_by_column: dict[int, str] = {}
    for index, header in enumerate(cells[0]):
        canonical = _COLUMNS.get(_normalize_header(header))
        if canonical is not None:
            field_by_column[index] = canonical
    missing = [
        canonical
        for canonical in _COLUMNS.values()
        if canonical not in field_by_column.values()
    ]
    if missing:
        raise BadTemplate(f"Missing required column(s): {', '.join(sorted(missing))}")
    raw_rows = []
    for row_number, row in enumerate(cells[1:], start=2):
        values = {
            field: (row[column].strip() if column < len(row) else "")
            for column, field in field_by_column.items()
        }
        raw_rows.append(_parse_row(row_number, values))
    return raw_rows


def _blank(value: str | None) -> str | None:
    value = (value or "").strip()
    return value or None


def _parse_amount(value: str) -> Decimal:
    """A positive euro amount from the file: "12.50", "12,50", "1.234,56",
    "€ 12,50" all parse; the template's amounts are always positive."""
    v = value.strip().replace("€", "").replace(" ", "")
    if not v:
        raise InvalidOperation()
    if "," in v and "." in v:
        # Whichever separator comes last is the decimal one: "1.234,56" and
        # "1,234.56" both become "1234.56".
        if v.rfind(",") > v.rfind("."):
            v = v.replace(".", "").replace(",", ".")
        else:
            v = v.replace(",", "")
    elif "," in v:
        v = v.replace(",", ".")
    return Decimal(v)


def _parse_location(value: str) -> tuple[str, str]:
    """The template's location column carries coordinates ("lat,lon" or
    "lat;lon"); the Maps link is built on the frontend and never stored as
    text (CONTEXT.md)."""
    parts = re.split(r"[,;]", value.strip())
    if len(parts) != 2:
        raise ValueError
    latitude = Decimal(parts[0].strip())
    longitude = Decimal(parts[1].strip())
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        raise ValueError
    lat, lon = fmt_coord(latitude), fmt_coord(longitude)
    assert lat is not None and lon is not None  # coordinates are never null here
    return lat, lon


def _parse_row(number: int, values: dict[str, str]) -> RawRow:
    raw = RawRow(
        number=number,
        type=_blank(values.get("type")),
        date=_blank(values.get("date")),
        wallet=_blank(values.get("wallet")),
        source_wallet=_blank(values.get("source_wallet")),
        destination_wallet=_blank(values.get("destination_wallet")),
        category=_blank(values.get("category")),
        description=_blank(values.get("description")),
    )
    errors: list[str] = []
    if raw.type is None or raw.type.lower() not in (
        TransactionType.EXPENSE.value,
        TransactionType.INCOME.value,
        TransactionType.TRANSFER.value,
    ):
        errors.append(
            "Type must be expense, income, or transfer"
            + (f" (got '{raw.type}')" if raw.type is not None else "")
        )
    else:
        raw.type = raw.type.lower()
    if raw.date is None:
        errors.append("Date is required")
    else:
        try:
            from_rome_day(raw.date)
        except ValueError:
            errors.append(f"Invalid date '{raw.date}' (use YYYY-MM-DD)")
    amount_text = values.get("amount")
    if amount_text is None or not amount_text.strip():
        errors.append("Amount is required")
    else:
        try:
            amount = _parse_amount(amount_text)
        except (InvalidOperation, ValueError):
            errors.append(f"Invalid amount '{amount_text.strip()}'")
        else:
            if amount <= 0:
                errors.append(f"Amount must be positive (got '{amount_text.strip()}')")
            elif amount > _MAX_AMOUNT:
                errors.append(f"Amount is too large (got '{amount_text.strip()}')")
            else:
                raw.amount = amount
    if raw.description is not None and len(raw.description) > 500:
        errors.append("Description is too long (max 500 characters)")
    location = values.get("location")
    if location is not None and location.strip():
        try:
            raw.latitude, raw.longitude = _parse_location(location)
        except (InvalidOperation, ValueError):
            errors.append(
                f"Invalid location '{location.strip()}' (use 'lat,lon')"
            )
    if errors:
        raw.error = "; ".join(errors)
    return raw


# --- Validation against the Account and the database -------------------------


def _resolve_wallet(session: Session, account_id: int, name: str) -> Wallet:
    """The Account's Wallet named `name` (unique per Account, case-insensitive),
    or an import error naming the missing or frozen Wallet. The frozen check is
    a resolution concern — the name resolved to a Wallet the import may not
    write to — and gives a message that names the Wallet; the service's own
    freeze rule still guards the write (validate_create / create_transaction)."""
    wallet = session.scalar(
        select(Wallet).where(
            Wallet.account_id == account_id,
            func.lower(Wallet.name) == name.lower(),
        )
    )
    if wallet is None:
        raise ImportValidationError(f"Unknown wallet '{name}'")
    if wallet.frozen:
        raise ImportValidationError(f"Wallet '{name}' is frozen")
    return wallet


def _resolve_category(
    session: Session, account_id: int, name: str | None, *, type: str
) -> Category | None:
    """The Account's Category named `name` with the given Type, or an import
    error. A blank name means no Category (the Transaction stays uncategorized).
    A name can exist for both Types (an expense "Food" and an income "Food"
    coexist), so the lookup filters by Type and only then falls back to the
    "wrong type" message."""
    if name is None:
        return None
    category = session.scalar(
        select(Category).where(
            Category.account_id == account_id,
            Category.type == type,
            func.lower(Category.name) == name.lower(),
        )
    )
    if category is not None:
        return category
    other_type = session.scalar(
        select(Category).where(
            Category.account_id == account_id,
            func.lower(Category.name) == name.lower(),
        )
    )
    if other_type is not None:
        raise ImportValidationError(
            f"Category '{name}' is a {other_type.type} category, not {type}"
        )
    raise ImportValidationError(f"Unknown {type} category '{name}'")


def _coord(value: str | None) -> Decimal | None:
    return Decimal(value) if value is not None else None


def _resolve_row(session: Session, account_id: int, raw: RawRow) -> dict[str, Any]:
    """Resolve the row's names to ids — the import's only job here. The
    CONTEXT.md rules themselves are enforced by `services.transactions`
    (validate_create / create_transaction) on the resolved ids, so a rule
    change lands in one place and an imported row cannot drift from a typed
    one. The guards below are template-shape rules — which COLUMNS a row of a
    given Type may fill in the file — and stay here because the service
    reasons in ids, not template columns.

    Raises ImportValidationError with the row's message; returns the kwargs
    for transaction_service.create_transaction / validate_create."""
    if raw.type == TransactionType.TRANSFER.value:
        if raw.wallet is not None:
            raise ImportValidationError(
                "Transfer rows use source wallet and destination wallet, not wallet"
            )
        if raw.category is not None:
            raise ImportValidationError("Transfers never carry a category")
        source_wallet_id = (
            _resolve_wallet(session, account_id, raw.source_wallet).id
            if raw.source_wallet is not None
            else None
        )
        destination_wallet_id = (
            _resolve_wallet(session, account_id, raw.destination_wallet).id
            if raw.destination_wallet is not None
            else None
        )
        return {
            "type": raw.type,
            "amount": raw.amount,
            "date": raw.date,
            "wallet_id": None,
            "source_wallet_id": source_wallet_id,
            "destination_wallet_id": destination_wallet_id,
            "category_id": None,
            "description": raw.description,
            "latitude": _coord(raw.latitude),
            "longitude": _coord(raw.longitude),
        }
    if raw.source_wallet is not None or raw.destination_wallet is not None:
        raise ImportValidationError(
            "source and destination wallets are only for Transfers"
        )
    wallet_id = None
    if raw.wallet is not None:
        wallet_id = _resolve_wallet(session, account_id, raw.wallet).id
    category_id = None
    if raw.category is not None and raw.type is not None:
        category = _resolve_category(
            session, account_id, raw.category, type=raw.type
        )
        assert category is not None  # a non-blank name either resolves or raises
        category_id = category.id
    return {
        "type": raw.type,
        "amount": raw.amount,
        "date": raw.date,
        "wallet_id": wallet_id,
        "source_wallet_id": None,
        "destination_wallet_id": None,
        "category_id": category_id,
        "description": raw.description,
        "latitude": _coord(raw.latitude),
        "longitude": _coord(raw.longitude),
    }


def _is_duplicate(
    session: Session, account_id: int, params: dict[str, Any]
) -> bool:
    """A row already in the database. Expense/income rows key on date + amount
    + type + wallet + category + description; transfer rows on date + amount +
    source + destination + description. The Type is part of the expense/income
    key — a deliberate deviation from the spec's literal key (US31/ID13), which
    lacks it and would make a €5 Expense and a €5 Income on the same Wallet and
    Category collide. The description is part of every key (ADR-0006), and a
    blank one matches a missing one (the column can hold either NULL or "").
    Opening Balance rows never match (they are not expense/income)."""
    day = from_rome_day(params["date"])
    stmt = (
        select(Transaction.id)
        .where(
            Transaction.account_id == account_id,
            Transaction.amount == params["amount"],
            Transaction.date >= day,
            Transaction.date < day + timedelta(days=1),
            func.coalesce(Transaction.description, "")
            == (params["description"] or ""),
        )
        .limit(1)
    )
    if params["type"] == TransactionType.TRANSFER.value:
        stmt = stmt.where(
            Transaction.type == TransactionType.TRANSFER.value,
            Transaction.source_wallet_id == params["source_wallet_id"],
            Transaction.destination_wallet_id == params["destination_wallet_id"],
        )
    else:
        stmt = stmt.where(
            Transaction.type == params["type"],
            Transaction.wallet_id == params["wallet_id"],
        )
        if params["category_id"] is None:
            stmt = stmt.where(Transaction.category_id.is_(None))
        else:
            stmt = stmt.where(Transaction.category_id == params["category_id"])
    return session.scalar(stmt) is not None


def _duplicate_key(params: dict[str, Any]) -> tuple:
    """The duplicate key, as a comparable tuple — keyed exactly as
    `_is_duplicate` queries it, including the Type dimension (the US31/ID13
    deviation rationale lives there) and the description (ADR-0006). The
    in-file `seen` set uses this so a row repeated in the file is flagged like
    one already in the database. The amount is quantized so "12.5" and
    "12.50" key identically; the description keys as "" when blank or
    missing."""
    amount = params["amount"].quantize(Decimal("0.01"))
    # The description keys as "" for both None and blank (ADR-0006): a blank
    # description matches a missing one in the file and in the database.
    description = params["description"] or ""
    if params["type"] == TransactionType.TRANSFER.value:
        return (
            "transfer",
            params["date"],
            amount,
            params["source_wallet_id"],
            params["destination_wallet_id"],
            description,
        )
    return (
        params["type"],
        params["date"],
        amount,
        params["wallet_id"],
        params["category_id"],
        description,
    )


# --- The two HTTP-facing operations ------------------------------------------


def preview_rows(
    session: Session, account_id: int, raw_rows: list[RawRow]
) -> list[ImportRow]:
    """Validate each parsed row against the Account and the database and return
    the preview verdicts. Nothing is inserted (US 31/32)."""
    rows: list[ImportRow] = []
    # Keys already seen in THIS file: a row repeated in the file is flagged
    # "duplicate" like a row already in the database, so the preview shows the
    # user what to drop before confirming (the insert itself would reject it).
    seen: set[tuple] = set()
    for raw in raw_rows:
        row = ImportRow(
            row=raw.number,
            type=raw.type,
            date=raw.date,
            amount=raw.amount,
            wallet=raw.wallet,
            source_wallet=raw.source_wallet,
            destination_wallet=raw.destination_wallet,
            category=raw.category,
            description=raw.description,
            latitude=raw.latitude,
            longitude=raw.longitude,
        )
        if raw.error is not None:
            row.status = "error"
            row.error = raw.error
        else:
            try:
                params = _resolve_row(session, account_id, raw)
                transaction_service.validate_create(
                    session,
                    account_id,
                    type=params["type"],
                    wallet_id=params["wallet_id"],
                    source_wallet_id=params["source_wallet_id"],
                    destination_wallet_id=params["destination_wallet_id"],
                    category_id=params["category_id"],
                )
            except (
                ImportValidationError,
                transaction_service.TransactionRuleError,
                scoping.NotOwned,
            ) as error:
                row.status = "error"
                row.error = str(error)
            else:
                key = _duplicate_key(params)
                if _is_duplicate(session, account_id, params) or key in seen:
                    row.status = "duplicate"
                else:
                    row.status = "ok"
                    seen.add(key)
        rows.append(row)
    return rows


def confirm_rows(
    session: Session, account_id: int, rows: list[ImportRowInput]
) -> list[Transaction]:
    """Insert the confirmed rows in one transaction (US 31/32): every row is
    re-validated (names re-resolved, rules re-checked, duplicates re-checked),
    and any failure rolls back the whole batch — nothing is written."""
    created: list[Transaction] = []
    try:
        for item in rows:
            raw = RawRow(
                number=item.row or 0,
                type=item.type,
                date=item.date,
                amount=item.amount,
                wallet=item.wallet,
                source_wallet=item.source_wallet,
                destination_wallet=item.destination_wallet,
                category=item.category,
                description=_blank(item.description),
                latitude=item.latitude,
                longitude=item.longitude,
            )
            params = _resolve_row(session, account_id, raw)
            if _is_duplicate(session, account_id, params):
                raise ImportValidationError(
                    f"Row {item.row or '?'} duplicates an existing transaction"
                )
            created.append(
                transaction_service.create_transaction(
                    session, account_id, commit=False, **params
                )
            )
        session.commit()
    except ImportValidationError:
        session.rollback()
        raise
    except (
        transaction_service.TransactionRuleError,
        scoping.NotOwned,
    ) as error:
        session.rollback()
        raise ImportValidationError(str(error)) from error
    for transaction in created:
        session.refresh(transaction)
    return created
