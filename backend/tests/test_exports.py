"""Excel Export (US 7.3): the pure workbook builder (unit seam) and the
endpoint (integration seam).

The unit tests feed hand-built ExportRows through `build_export_workbook` and
re-open the produced bytes with openpyxl, asserting the template's cells
literally; the integration tests drive GET /transactions/export through the
HTTP seam against the seeded database.
"""

import re
from decimal import Decimal
from io import BytesIO
from itertools import count

from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db import create_db_engine
from app.models import Wallet, WalletType
from app.services.exports import ExportRow, build_export_workbook
from conftest import (
    SEED_EMAIL,
    SEED_PASSWORD,
    insert_foreign_account,
)

# The database resets once per session, so entity names (unique per Account)
# must not repeat across tests; a per-process counter keeps them distinct.
_UNIQUE = count(1)


def _name(prefix: str) -> str:
    return f"{prefix} {next(_UNIQUE)}"


HEADER = [
    "date",
    "type",
    "amount",
    "wallet",
    "source wallet",
    "destination wallet",
    "category",
    "description",
    "location",
]


def _cells(content: bytes) -> list[list[str]]:
    """The sheet's rows as plain strings — None cells become empty strings —
    so an assertion reads like the template's own cells."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [
            ["" if cell is None else str(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


# --- The pure builder --------------------------------------------------------


def test_build_export_workbook_writes_only_the_header_for_no_rows() -> None:
    content = build_export_workbook([])
    assert _cells(content) == [HEADER]


def test_build_export_workbook_writes_an_expense_row_literally() -> None:
    content = build_export_workbook(
        [
            ExportRow(
                date="2026-08-01",
                type="expense",
                amount=Decimal("12.50"),
                wallet="Checking",
                source_wallet=None,
                destination_wallet=None,
                category="Food",
                description="Lunch",
                latitude=Decimal("45.4642"),
                longitude=Decimal("9.1900"),
            ),
        ]
    )
    assert _cells(content) == [
        HEADER,
        ["2026-08-01", "expense", "12.50", "Checking", "", "", "Food", "Lunch", "45.4642,9.19"],
    ]


def test_build_export_workbook_writes_a_transfer_row() -> None:
    content = build_export_workbook(
        [
            ExportRow(
                date="2026-08-02",
                type="transfer",
                amount=Decimal("50"),
                wallet=None,
                source_wallet="Checking",
                destination_wallet="Savings",
                category=None,
                description=None,
                latitude=None,
                longitude=None,
            ),
        ]
    )
    assert _cells(content) == [
        HEADER,
        ["2026-08-02", "transfer", "50.00", "", "Checking", "Savings", "", "", ""],
    ]


def test_build_export_workbook_formats_amounts_to_two_decimals() -> None:
    content = build_export_workbook(
        [
            ExportRow(
                date="2026-08-03",
                type="income",
                amount=Decimal("100"),
                wallet="Checking",
                source_wallet=None,
                destination_wallet=None,
                category=None,
                description=None,
                latitude=None,
                longitude=None,
            ),
            ExportRow(
                date="2026-08-04",
                type="expense",
                amount=Decimal("0.01"),
                wallet="Checking",
                source_wallet=None,
                destination_wallet=None,
                category=None,
                description="",
                latitude=None,
                longitude=None,
            ),
        ]
    )
    assert [row[2] for row in _cells(content)[1:]] == ["100.00", "0.01"]
    assert [row[7] for row in _cells(content)[1:]] == ["", ""]


# --- The endpoint: GET /transactions/export through the HTTP seam ------------


async def _login(client: AsyncClient) -> str:
    response = await client.post(
        "/auth/login", json={"email": SEED_EMAIL, "password": SEED_PASSWORD}
    )
    assert response.status_code == 200
    return response.json()["access_token"]


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def _create_wallet(
    client: AsyncClient, token: str, name: str, opening_balance: str = "0.00"
) -> int:
    response = await client.post(
        "/wallets",
        json={"name": name, "type": "checking", "opening_balance": opening_balance},
        headers=_auth(token),
    )
    assert response.status_code == 201, response.text
    return response.json()["id"]


async def _create_category(
    client: AsyncClient, token: str, name: str, type: str
) -> int:
    response = await client.post(
        "/categories",
        json={"name": name, "type": type, "color": "#ef4444"},
        headers=_auth(token),
    )
    assert response.status_code == 201, response.text
    return response.json()["id"]


async def _create_transaction(client: AsyncClient, token: str, **payload: object) -> None:
    response = await client.post(
        "/transactions", json=payload, headers=_auth(token)
    )
    assert response.status_code == 201, response.text


async def _export(client: AsyncClient, token: str, **params: str | int) -> bytes:
    response = await client.get(
        "/transactions/export", params=params, headers=_auth(token)
    )
    assert response.status_code == 200, response.text
    return response.content


async def _seed_ledger(
    client: AsyncClient, token: str
) -> dict[str, int | str]:
    """Two Wallets, an expense Category and an income Category, and five
    Transactions spread across dates, Wallets, Categories, and descriptions
    (one Transfer between the Wallets). Returns the created ids and names,
    so assertions can name the rows they expect. Tests that assert exact
    row sets filter by their own Wallet (the database accumulates across
    the session), so the seed's fixed dates cannot collide with other
    tests' rows."""
    wallet_a_name = _name("Export A")
    wallet_b_name = _name("Export B")
    wallet_a = await _create_wallet(client, token, wallet_a_name)
    wallet_b = await _create_wallet(client, token, wallet_b_name)
    food_name = _name("Export Food")
    food = await _create_category(client, token, food_name, "expense")
    salary_name = _name("Export Salary")
    salary = await _create_category(client, token, salary_name, "income")
    await _create_transaction(
        client, token,
        type="expense", amount="10.00", date="2026-07-10",
        wallet_id=wallet_a, category_id=food, description="coffee",
    )
    await _create_transaction(
        client, token,
        type="expense", amount="12.50", date="2026-08-01",
        wallet_id=wallet_a, category_id=food, description="lunch",
        latitude="45.4642", longitude="9.1900",
    )
    await _create_transaction(
        client, token,
        type="expense", amount="30.00", date="2026-08-15",
        wallet_id=wallet_b, category_id=food, description="dinner",
    )
    await _create_transaction(
        client, token,
        type="income", amount="200.00", date="2026-08-20",
        wallet_id=wallet_a, category_id=salary, description="salary",
    )
    await _create_transaction(
        client, token,
        type="transfer", amount="50.00", date="2026-08-25",
        source_wallet_id=wallet_a, destination_wallet_id=wallet_b,
        description="savings transfer",
    )
    return {
        "wallet_a": wallet_a,
        "wallet_b": wallet_b,
        "food": food,
        "salary": salary,
        "wallet_a_name": wallet_a_name,
        "wallet_b_name": wallet_b_name,
        "food_name": food_name,
        "salary_name": salary_name,
    }


async def test_export_leaves_out_opening_balance_rows(client: AsyncClient) -> None:
    token = await _login(client)
    # A nonzero opening balance seeds an Opening Balance Transaction, which
    # the template's type vocabulary cannot carry (ADR-0015): a Wallet whose
    # only Transaction is its opening balance exports an empty sheet.
    seeds = await _create_wallet(client, token, _name("Export Seeds"), "100.00")

    rows = _cells(await _export(client, token, wallet_id=seeds))

    assert rows == [HEADER]


async def test_export_writes_the_ledger_with_names_and_shapes(
    client: AsyncClient,
) -> None:
    token = await _login(client)
    data = await _seed_ledger(client, token)

    rows = _cells(await _export(client, token, wallet_id=data["wallet_a"]))

    # Date-ascending, names resolved, amounts two-decimal text, the location
    # column as "lat,lon" when coordinates exist and blank when they do not.
    assert rows == [
        HEADER,
        [
            "2026-07-10", "expense", "10.00", str(data["wallet_a_name"]), "", "",
            str(data["food_name"]), "coffee", "",
        ],
        [
            "2026-08-01", "expense", "12.50", str(data["wallet_a_name"]), "", "",
            str(data["food_name"]), "lunch", "45.4642,9.19",
        ],
        [
            "2026-08-20", "income", "200.00", str(data["wallet_a_name"]), "", "",
            str(data["salary_name"]), "salary", "",
        ],
        # The transfer row uses source/destination and never wallet/category.
        [
            "2026-08-25", "transfer", "50.00", "", str(data["wallet_a_name"]),
            str(data["wallet_b_name"]), "", "savings transfer", "",
        ],
    ]


async def test_export_applies_the_wallet_filter(client: AsyncClient) -> None:
    token = await _login(client)
    data = await _seed_ledger(client, token)

    rows = _cells(await _export(client, token, wallet_id=data["wallet_b"]))

    # Only B's own expense and the Transfer touching B on either leg.
    assert [row[1] for row in rows[1:]] == ["expense", "transfer"]
    assert [row[0] for row in rows[1:]] == ["2026-08-15", "2026-08-25"]


async def test_export_applies_the_category_filter(client: AsyncClient) -> None:
    token = await _login(client)
    data = await _seed_ledger(client, token)

    rows = _cells(await _export(client, token, category_id=data["food"]))

    # The Category is this test's own (no other test shares its id), so the
    # rows are exactly the seed's three food expenses.
    assert [row[1] for row in rows[1:]] == ["expense", "expense", "expense"]
    assert [row[0] for row in rows[1:]] == ["2026-07-10", "2026-08-01", "2026-08-15"]


async def test_export_applies_the_date_range_filter(client: AsyncClient) -> None:
    token = await _login(client)
    wallet = await _create_wallet(client, token, _name("Export Range"))
    for day in ("2027-03-01", "2027-03-10", "2027-03-20"):
        await _create_transaction(
            client, token,
            type="expense", amount="5.00", date=day, wallet_id=wallet,
        )

    rows = _cells(
        await _export(
            client, token,
            wallet_id=wallet, from_date="2027-03-05", to_date="2027-03-15",
        )
    )

    assert [row[0] for row in rows[1:]] == ["2027-03-10"]


async def test_export_applies_the_search_filter(client: AsyncClient) -> None:
    token = await _login(client)
    data = await _seed_ledger(client, token)

    rows = _cells(
        await _export(client, token, wallet_id=data["wallet_a"], q="lunch")
    )

    assert [row[0] for row in rows[1:]] == ["2026-08-01"]


async def test_export_includes_rows_on_frozen_wallets(client: AsyncClient) -> None:
    token = await _login(client)
    a_name = _name("Export A")
    frozen_name = _name("Export Frozen")
    wallet_a = await _create_wallet(client, token, a_name)
    frozen = await _create_wallet(client, token, frozen_name)
    await _create_transaction(
        client, token,
        type="transfer", amount="100.00", date="2026-08-01",
        source_wallet_id=wallet_a, destination_wallet_id=frozen,
    )
    await _create_transaction(
        client, token,
        type="transfer", amount="100.00", date="2026-08-02",
        source_wallet_id=frozen, destination_wallet_id=wallet_a,
    )
    # Freeze the Wallet at Balance exactly €0 (ADR-0002).
    response = await client.delete(f"/wallets/{frozen}", headers=_auth(token))
    assert response.status_code == 204

    rows = _cells(await _export(client, token, wallet_id=frozen))

    # Frozen Wallets are read-only for writes; their history still exports.
    assert rows[1:] == [
        ["2026-08-01", "transfer", "100.00", "", a_name, frozen_name, "", "", ""],
        ["2026-08-02", "transfer", "100.00", "", frozen_name, a_name, "", "", ""],
    ]


async def test_export_requires_auth(client: AsyncClient) -> None:
    response = await client.get("/transactions/export")
    assert response.status_code == 401


async def test_export_scopes_foreign_wallets(
    client: AsyncClient, database_url: str
) -> None:
    token = await _login(client)
    account_id = insert_foreign_account(database_url, "export-thief@budjetame.dev")
    engine = create_db_engine(database_url)
    try:
        with Session(engine) as session:
            wallet = Wallet(
                account_id=account_id,
                name="Foreign Funds",
                type=WalletType.CHECKING.value,
            )
            session.add(wallet)
            session.commit()
            wallet_id = wallet.id
    finally:
        engine.dispose()

    response = await client.get(
        "/transactions/export",
        params={"wallet_id": wallet_id},
        headers=_auth(token),
    )

    assert response.status_code == 403


async def test_export_has_attachment_headers(client: AsyncClient) -> None:
    token = await _login(client)
    await _seed_ledger(client, token)

    response = await client.get("/transactions/export", headers=_auth(token))

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = response.headers["content-disposition"]
    assert disposition.startswith('attachment; filename="')
    assert re.search(r'budjetame-\d{4}-\d{2}-\d{2}\.xlsx"$', disposition)
